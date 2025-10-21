# extract_wbs_json_v7.py
# Usage: python extract_wbs_json_v7.py Book1.xlsx --out wbs_all.json
from __future__ import annotations
from pathlib import Path
from typing import List, Dict, Any, Tuple
import argparse, json, re
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, date

REQUIRED_COLS = [
    "Planned Finish", "Forecast Finish", "Schedule %", "Earned %",
    "ecart", "impact", "Glissement"
]

# ---------- Helpers (format identique à ton exemple) ----------
def as_text(v: Any) -> str:
    """Dates -> 'dd-Mon-yy', sinon str, None -> '' """
    if isinstance(v, (datetime, date)):
        return v.strftime("%d-%b-%y")
    return "" if v is None else str(v)

def parse_percent_float(v: Any) -> float:
    """
    Pour Schedule/Earned: retourne un float 0–100 avec 2 décimales max
    (ex: '75.51%' -> 75.51, 0.7551 -> 75.51)
    """
    if v is None or v == "":
        return 0.0
    if isinstance(v, str):
        s = v.strip()
        s = re.sub(r"[^\d\-\.,%]", "", s).replace(",", ".")
        s = s.replace("%", "")
        if s in ("", "-", "."):
            return 0.0
        try:
            val = float(s)
        except Exception:
            return 0.0
    else:
        try:
            val = float(v)
        except Exception:
            return 0.0
    if abs(val) <= 1.5:  # Excel peut stocker 75.51% sous 0.7551
        val *= 100.0
    # arrondi technique, puis si c'est un entier, on gardera un int via tidy_num()
    return round(val, 2)

def parse_percent_int(v: Any) -> int:
    """
    Pour ecart/impact: retourne un ENTIER en 0–100 (ex: '-4%' -> -4, 0.05 -> 5).
    """
    f = parse_percent_float(v)
    return int(round(f))  # force entier comme ton exemple

def tidy_num(x: float | int) -> float | int:
    """
    Pour schedule/earned: si 75.0 -> 75 (int), sinon 36.81 reste 36.81.
    """
    if isinstance(x, (int,)):
        return x
    if abs(x - round(x)) < 1e-9:
        return int(round(x))
    return x

def trim_empty_border(df: pd.DataFrame) -> pd.DataFrame:
    while df.shape[1] > 0 and df.iloc[:, 0].isna().all(): df = df.iloc[:, 1:]
    while df.shape[1] > 0 and df.iloc[:, -1].isna().all(): df = df.iloc[:, :-1]
    while df.shape[0] > 0 and df.iloc[0].isna().all(): df = df.iloc[1:, :]
    while df.shape[0] > 0 and df.iloc[-1].isna().all(): df = df.iloc[:-1, :]
    return df

def make_unique_columns(cols: List[str]) -> List[str]:
    seen: Dict[str,int] = {}; out: List[str] = []
    for c in cols:
        c = "" if c is None else str(c).strip()
        if c not in seen:
            seen[c] = 1; out.append(c)
        else:
            seen[c] += 1; out.append(f"{c}_{seen[c]}")
    return out

def leading_spaces(s: Any) -> int:
    if s is None: return 0
    s = str(s).replace("\t","    ")
    m = re.match(r"^\s*", s)
    return len(m.group(0)) if m else 0

def clean_label(s: Any) -> str:
    if s is None: return ""
    return str(s).strip()

def _norm(x: Any) -> str:
    return re.sub(r"\s+", " ", str(x or "")).strip().lower()

def has_all_required(headers: List[Any]) -> bool:
    H = [_norm(h) for h in headers]
    R = [_norm(rc) for rc in REQUIRED_COLS]
    return all(rc in H for rc in R)

def _is_date_like(x: str) -> bool:
    x = x.strip()
    if not x: return False
    if re.match(r"\d{4}-\d{2}-\d{2}", x): return True
    if re.match(r"\d{1,2}-[A-Za-z]{3}-\d{2}", x): return True
    try:
        datetime.strptime(x, "%d-%b-%y"); return True
    except Exception:
        return False

# ---------- Détection de tous les blocs (avec extension gauche pour colonne label) ----------
def detect_all_blocks_with_left_extension(ws, max_added_left: int = 5) -> List[Tuple[int,int,int,int]]:
    max_r, max_c = ws.max_row, ws.max_column
    blocks: List[Tuple[int,int,int,int]] = []
    r = 1
    while r <= max_r:
        headers = [ws.cell(r,c).value for c in range(1,max_c+1)]
        if not any(headers):
            r += 1; continue
        if not has_all_required(headers):
            r += 1; continue

        nz = [i+1 for i,v in enumerate(headers) if v not in (None,""," ")]
        c1, c2 = min(nz), max(nz)

        # descendre pour fin du bloc
        r2 = r + 1
        while r2 <= max_r:
            row_vals = [ws.cell(r2, c).value for c in range(c1, c2+1)]
            if all(v in (None, "", " ") for v in row_vals):
                break
            r2 += 1

        # extension gauche si colonne sans titre avec données
        added = 0
        while c1 > 1 and added < max_added_left:
            col_vals = [ws.cell(rr, c1-1).value for rr in range(r+1, r2)]
            if any(v not in (None,""," ") for v in col_vals):
                c1 -= 1
                added += 1
            else:
                break

        blocks.append((r, c1, r2-1, c2))
        # saute au-delà de ce bloc
        r = r2 + 1
    return blocks

# ---------- Choix de la colonne Label ----------
def pick_label_col(df: pd.DataFrame) -> str:
    """
    Priorité :
      1) Colonne d'en-tête vide "" si surtout du texte non date-like
      2) Sinon meilleure colonne non-requise : textuelle, non date-like, variété d'indentation
      3) Sinon première colonne
    """
    req = {"Planned Finish","Forecast Finish","Schedule %","Earned %","ecart","impact","Glissement"}
    if "" in df.columns:
        col = df[""].astype(str)
        if (col.str.strip() != "").any():
            not_dates_ratio = (~col.str.strip().apply(_is_date_like)).mean()
            if not_dates_ratio > 0.6:
                return ""

    best, best_score = None, -1.0
    for c in df.columns:
        if c in req: 
            continue
        col = df[c].astype(str)
        texty = col.apply(lambda s: any(ch.isalpha() for ch in s)).mean()
        not_date = col.apply(lambda s: not _is_date_like(s)).mean()
        indents = col.apply(leading_spaces)
        indent_levels = indents.nunique()
        score = (texty*2.0) + (not_date*1.5) + (min(indent_levels, 4)*0.8)
        if score > best_score:
            best_score, best = score, c
    return best or df.columns[0]

# ---------- WBS builder ----------
def to_wbs_tree(df: pd.DataFrame, label_col: str) -> Dict:
    df = df.copy()
    df[label_col] = df[label_col].fillna("")
    df = df[df[label_col].astype(str).str.strip() != ""]
    df["_indent"] = df[label_col].apply(leading_spaces)

    uniq = sorted(df["_indent"].unique().tolist()) or [0]
    uniq = uniq[:3]  # max 3 niveaux
    space2lvl = {sp: i+1 for i, sp in enumerate(uniq)}

    def row_metrics(r: pd.Series) -> dict:
        # NOTE: ecart/impact -> ENTIER, schedule/earned -> tidy (int si rond)
        schedule = tidy_num(parse_percent_float(r.get("Schedule %")))
        earned   = tidy_num(parse_percent_float(r.get("Earned %")))
        return {
            "planned_finish": as_text(r.get("Planned Finish")),
            "forecast_finish": as_text(r.get("Forecast Finish")),
            "schedule": schedule,
            "earned":   earned,
            "ecart":    parse_percent_int(r.get("ecart")),
            "impact":   parse_percent_int(r.get("impact")),
            "glissement": as_text(r.get("Glissement")),
        }

    root: Dict | None = None
    stack: List[Dict] = []

    for _, r in df.iterrows():
        label = clean_label(r[label_col])
        if not label:
            continue
        lvl = space2lvl.get(r["_indent"], len(space2lvl))  # fallback = plus profond
        node = {"label": label, "level": int(lvl), "metrics": row_metrics(r), "children": []}

        if not stack:
            root = node
            stack = [node]
            continue

        while stack and stack[-1]["level"] >= lvl:
            stack.pop()

        if not stack:
            # nouveau L1 → frère du root
            if root is None:
                root = node
                stack = [node]
            else:
                root.setdefault("children", []).append(node)
                stack = [root, node]
        else:
            stack[-1]["children"].append(node)
            stack.append(node)

    return root or {}

# ---------- Extraction (tous les tableaux) ----------
def extract_all_wbs(input_xlsx: str) -> List[Dict]:
    wb = load_workbook(input_xlsx, data_only=True)
    results: List[Dict] = []

    for ws in wb.worksheets:
        blocks = detect_all_blocks_with_left_extension(ws)
        for (r1, c1, r2, c2) in blocks:
            data = [[ws.cell(rr, cc).value for cc in range(c1, c2+1)] for rr in range(r1, r2+1)]
            if len(data) < 2:
                continue
            df = pd.DataFrame(data)
            df.columns = make_unique_columns([str(x).strip() if x is not None else "" for x in df.iloc[0].tolist()])
            df = df.iloc[1:, :].reset_index(drop=True)
            df = trim_empty_border(df)

            # Filtre: doit contenir toutes les colonnes requises
            if not all(col in df.columns for col in REQUIRED_COLS):
                continue

            label_col = pick_label_col(df)
            tree = to_wbs_tree(df, label_col)
            if tree:
                results.append({
                    "sheet": ws.title,
                    "range": f"R{r1}C{c1}:R{r2}C{c2}",
                    "wbs": tree
                })
    return results

# ---------- CLI ----------
if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Build an array of WBS JSONs from all valid tables in Excel.")
    p.add_argument("input_xlsx", help="Path to Excel file, e.g., Book1.xlsx")
    p.add_argument("--out", default="wbs_all.json", help="Output JSON path")
    args = p.parse_args()

    all_wbs = extract_all_wbs(args.input_xlsx)
    Path(args.out).write_text(json.dumps(all_wbs, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"Saved: {args.out}  |  Tables matched: {len(all_wbs)}")
