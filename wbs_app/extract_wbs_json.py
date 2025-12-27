# extract_wbs_json_v7.py
# Usage: python extract_wbs_json_v7.py Book1.xlsx --out wbs_all.json
from __future__ import annotations
from pathlib import Path
from typing import List, Dict, Any, Tuple
import argparse, json, re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from typing import Any

REQUIRED_COLS = [
    "Planned Finish", "Forecast Finish", "Schedule %", "Earned %",
    "ecart", "impact", "Glissement"
]

SUMMARY_HEADER_GROUPS = {
    "activity id": ["Activity ID", "ActivityID"],
    "bl project finish": [
        "BL Project Finish", "Baseline Project Finish",
        "BL Project Finish Date", "Baseline Project Finish Date"
    ],
    "finish": ["Finish", "Project Finish", "Finish Date"],
    "units % complete": ["Units % Complete", "Units Percent Complete", "% Complete", "Percent Complete"],
    "variance - bl project finish date": [
        "Variance - BL Project Finish Date", "Variance BL Project Finish Date",
        "Variance - BL Project Finish", "Variance BL Project Finish"
    ],
}

ASSIGN_HEADER_GROUPS = {
    "activity id": ["Activity ID", "ActivityID"],
    "activity name": ["Activity Name", "ActivityName"],
    "start": ["Start", "Start Date"],
    "finish": ["Finish", "Finish Date"],
    "rcv / phases": ["RCV / Phases", "RCV/Phases", "RCV Phases"],
    "budgeted units": ["Budgeted Units", "Budget Units"],
    "spreadsheet field": ["Spreadsheet Field", "SpreadsheetField"],
}

SUMMARY_REQUIRED_FIELDS = [
    "Activity ID",
    "BL Project Finish",
    "Finish",
    "Units % Complete",
    "Variance - BL Project Finish Date",
]
SUMMARY_OPTIONAL_FIELDS = [
    "Activity Name",
    "Activity Status",
    "Budgeted Labor Units",
]
ASSIGN_REQUIRED_FIELDS = [
    "Activity ID",
    "Budgeted Units",
    "Spreadsheet Field",
]
ASSIGN_OPTIONAL_FIELDS = [
    "Activity Name",
    "Start",
    "Finish",
    "RCV / Phases",
]

SUMMARY_FIELD_VARIANTS = {
    "Activity ID": SUMMARY_HEADER_GROUPS["activity id"] + ["Activity ID"],
    "BL Project Finish": SUMMARY_HEADER_GROUPS["bl project finish"] + ["BL Project Finish"],
    "Finish": SUMMARY_HEADER_GROUPS["finish"] + ["Finish"],
    "Units % Complete": SUMMARY_HEADER_GROUPS["units % complete"] + ["Units % Complete"],
    "Variance - BL Project Finish Date": SUMMARY_HEADER_GROUPS["variance - bl project finish date"]
    + ["Variance - BL Project Finish Date"],
    "Activity Name": ["Activity Name", "ActivityName"],
    "Activity Status": ["Activity Status", "Status"],
    "Budgeted Labor Units": ["Budgeted Labor Units", "Budgeted Units"],
}

ASSIGN_FIELD_VARIANTS = {
    "Activity ID": ASSIGN_HEADER_GROUPS["activity id"] + ["Activity ID"],
    "Activity Name": ASSIGN_HEADER_GROUPS["activity name"] + ["Activity Name"],
    "Start": ASSIGN_HEADER_GROUPS["start"] + ["Start"],
    "Finish": ASSIGN_HEADER_GROUPS["finish"] + ["Finish"],
    "RCV / Phases": ASSIGN_HEADER_GROUPS["rcv / phases"] + ["RCV / Phases"],
    "Budgeted Units": ASSIGN_HEADER_GROUPS["budgeted units"] + ["Budgeted Units"],
    "Spreadsheet Field": ASSIGN_HEADER_GROUPS["spreadsheet field"] + ["Spreadsheet Field"],
}

# ---------- Helpers (format identique à ton exemple) ----------
def as_text(v: Any) -> str:
    """Dates -> 'dd-Mon-yy', sinon str, None -> '' """
    if v is None or str(v).strip() == "":
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%d-%b-%y")
    d = _to_excel_date(v)
    if d:
        return d.strftime("%d-%b-%y")
    return str(v)



def parse_percent_float(v):
    """
    Lecture propre des pourcentages :
    - '75.51%' -> 75.51
    - 0.7551   -> 75.51
    - '0.7551' -> 75.51
    - '1.2%'   -> 1.2
    - 75.51    -> 75.51
    """
    import re
    if v is None or v == "":
        return 0.0
    if isinstance(v, str):
        s = re.sub(r"[^\d\-\.,%]", "", v).replace(",", ".").replace("%", "").strip()
        if s in ("", "-", "."):
            return 0.0
        try:
            val = float(s)
        except Exception:
            return 0.0
    else:
        try:
            val = float(v)
        except Exception:
            return 0.0

    # ✅ multiplie uniquement si la valeur est comprise entre -1 et 1
    if -1.0 <= val <= 1.0:
        val *= 100.0

    return val




def parse_percent_int(v: Any) -> int:
    """
    Pour ecart/impact: retourne un ENTIER en 0–100 (ex: '-4%' -> -4, 0.05 -> 5).
    """
    f = parse_percent_float(v)
    return int(round(f))  # force entier comme ton exemple

def tidy_num(x: float | int) -> float | int:
    """
    Pour schedule/earned: si 75.0 -> 75 (int), sinon 36.81 reste 36.81.
    """
    if isinstance(x, (int,)):
        return x
    if abs(x - round(x)) < 1e-9:
        return int(round(x))
    return x

def _safe_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    try:
        return float(v)
    except Exception:
        return None

def _parse_days(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).lower().replace("j", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None

def _to_excel_date(v: Any) -> date | None:
    try:
        from openpyxl.utils.datetime import from_excel
    except Exception:
        from_excel = None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)) and from_excel is not None:
        try:
            return from_excel(v).date()
        except Exception:
            return None
    s = str(v or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%b-%y", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None

def _week_start(d: date) -> date:
    from datetime import timedelta
    return d - timedelta(days=d.weekday())

def _find_header_idx(headers: list[Any], name: str) -> int | None:
    target = str(name).strip().lower()
    for idx, v in enumerate(headers):
        if str(v or "").strip().lower() == target:
            return idx
    return None

def trim_empty_border(df: pd.DataFrame) -> pd.DataFrame:
    while df.shape[1] > 0 and df.iloc[:, 0].isna().all(): df = df.iloc[:, 1:]
    while df.shape[1] > 0 and df.iloc[:, -1].isna().all(): df = df.iloc[:, :-1]
    while df.shape[0] > 0 and df.iloc[0].isna().all(): df = df.iloc[1:, :]
    while df.shape[0] > 0 and df.iloc[-1].isna().all(): df = df.iloc[:-1, :]
    return df

def trim_empty_border_with_offsets(df: pd.DataFrame) -> Tuple[pd.DataFrame, int, int, int, int]:
    top = left = bottom = right = 0
    while df.shape[1] > 0 and df.iloc[:, 0].isna().all():
        df = df.iloc[:, 1:]
        left += 1
    while df.shape[1] > 0 and df.iloc[:, -1].isna().all():
        df = df.iloc[:, :-1]
        right += 1
    while df.shape[0] > 0 and df.iloc[0].isna().all():
        df = df.iloc[1:, :]
        top += 1
    while df.shape[0] > 0 and df.iloc[-1].isna().all():
        df = df.iloc[:-1, :]
        bottom += 1
    return df, top, left, bottom, right

def _sheet_ref(sheet: str) -> str:
    if re.search(r"[^A-Za-z0-9_]", sheet):
        return "'" + sheet.replace("'", "''") + "'"
    return sheet

def _cell_ref(meta: Dict[str, Any] | None, row_idx: int | None, col_idx: int | None) -> str | None:
    if meta is None or row_idx is None or col_idx is None:
        return None
    sheet = meta.get("sheet")
    row_base = meta.get("data_row_start")
    col_base = meta.get("data_col_start")
    if sheet is None or row_base is None or col_base is None:
        return None
    row_num = row_base + row_idx
    col_num = col_base + col_idx
    if row_num <= 0 or col_num <= 0:
        return None
    return f"{_sheet_ref(sheet)}!{get_column_letter(col_num)}{row_num}"

def _append_tip_sources(tip: str | None, sources: List[str], prefix: str = "Cells") -> str | None:
    items = [s for s in sources if s]
    if not items:
        return tip
    line = f"{prefix}:\n" + "\n".join(f"- {s}" for s in items)
    if tip:
        return f"{tip}\n{line}"
    return line

def make_unique_columns(cols: List[str]) -> List[str]:
    seen: Dict[str,int] = {}; out: List[str] = []
    for c in cols:
        c = "" if c is None else str(c).strip()
        if c not in seen:
            seen[c] = 1; out.append(c)
        else:
            seen[c] += 1; out.append(f"{c}_{seen[c]}")
    return out

def _apply_column_mapping(
    df: pd.DataFrame,
    raw_headers: list[Any],
    mapping: dict[str, str] | None,
) -> Tuple[pd.DataFrame, list[Any]]:
    if not mapping:
        return df, raw_headers
    actual_to_canonical: Dict[str, str] = {}
    for canonical, actual in mapping.items():
        if actual:
            actual_to_canonical[str(actual).strip()] = canonical
    if not actual_to_canonical:
        return df, raw_headers

    new_headers: list[Any] = []
    for h in raw_headers:
        key = str(h).strip()
        new_headers.append(actual_to_canonical.get(key, h))

    rename_map: Dict[str, str] = {}
    for col in df.columns:
        key = str(col).strip()
        canonical = actual_to_canonical.get(key)
        if canonical:
            rename_map[col] = canonical
    if rename_map:
        df = df.rename(columns=rename_map)
    return df, new_headers

def _table_field_variants(table_type: str) -> Dict[str, List[str]]:
    if table_type == "activity_summary":
        return SUMMARY_FIELD_VARIANTS
    if table_type == "resource_assignments":
        return ASSIGN_FIELD_VARIANTS
    return {}

def suggest_column_mapping(raw_headers: list[Any], table_type: str) -> Dict[str, str]:
    variants = _table_field_variants(table_type)
    mapping: Dict[str, str] = {}
    for canonical, names in variants.items():
        idx = _find_header_idx_norm(raw_headers, names)
        if idx is not None:
            mapping[canonical] = str(raw_headers[idx]).strip()
    return mapping

def get_table_headers(
    input_xlsx: str,
    table_type: str,
    column_mapping: dict[str, dict[str, str]] | None = None,
) -> Tuple[list[Any], Dict[str, Any]] | None:
    table = _load_detected_table(input_xlsx, table_type, column_mapping=column_mapping)
    if not table:
        return None
    _, meta, raw_headers = table
    return raw_headers, meta

def leading_spaces(s: Any) -> int:
    if s is None: return 0
    s = str(s).replace("\t","    ")
    m = re.match(r"^\s*", s)
    return len(m.group(0)) if m else 0

def clean_label(s: Any) -> str:
    if s is None: return ""
    return str(s).strip()

def _norm(x: Any) -> str:
    return re.sub(r"\s+", " ", str(x or "")).strip().lower()

def _norm_header(x: Any) -> str:
    s = str(x or "").strip().lower()
    if not s:
        return ""
    s = s.replace("%", " percent ")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def has_all_required(headers: List[Any]) -> bool:
    H = [_norm(h) for h in headers]
    R = [_norm(rc) for rc in REQUIRED_COLS]
    return all(rc in H for rc in R)

def _is_date_like(x: str) -> bool:
    x = x.strip()
    if not x: return False
    if re.match(r"\d{4}-\d{2}-\d{2}", x): return True
    if re.match(r"\d{1,2}-[A-Za-z]{3}-\d{2}", x): return True
    try:
        datetime.strptime(x, "%d-%b-%y"); return True
    except Exception:
        return False

def _is_week_header(v: Any) -> bool:
    if isinstance(v, (datetime, date)):
        return True
    s = str(v or "").strip()
    if not s:
        return False
    if re.match(r"\d{4}-\d{2}-\d{2}", s):
        return True
    if re.match(r"\d{1,2}-[A-Za-z]{3}-\d{2}", s):
        return True
    if re.match(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", s):
        return True
    return False

def _match_header_groups(headers: List[Any], groups: dict) -> Tuple[List[str], List[str]]:
    header_set = { _norm_header(h) for h in headers if _norm_header(h) }
    matched = []
    missing = []
    for key, opts in groups.items():
        norm_opts = [_norm_header(o) for o in opts]
        if any(opt in header_set for opt in norm_opts):
            matched.append(key)
        else:
            missing.append(key)
    return matched, missing

# ---------- Détection de tous les blocs (avec extension gauche pour colonne label) ----------
def detect_all_blocks_with_left_extension(ws, max_added_left: int = 5) -> List[Tuple[int,int,int,int]]:
    max_r, max_c = ws.max_row, ws.max_column
    blocks: List[Tuple[int,int,int,int]] = []
    r = 1
    while r <= max_r:
        headers = [ws.cell(r,c).value for c in range(1,max_c+1)]
        if not any(headers):
            r += 1; continue
        if not has_all_required(headers):
            r += 1; continue

        nz = [i+1 for i,v in enumerate(headers) if v not in (None,""," ")]
        c1, c2 = min(nz), max(nz)

        # descendre pour fin du bloc
        r2 = r + 1
        while r2 <= max_r:
            row_vals = [ws.cell(r2, c).value for c in range(c1, c2+1)]
            if all(v in (None, "", " ") for v in row_vals):
                break
            r2 += 1

        # extension gauche si colonne sans titre avec données
        added = 0
        while c1 > 1 and added < max_added_left:
            col_vals = [ws.cell(rr, c1-1).value for rr in range(r+1, r2)]
            if any(v not in (None,""," ") for v in col_vals):
                c1 -= 1
                added += 1
            else:
                break

        blocks.append((r, c1, r2-1, c2))
        # saute au-delà de ce bloc
        r = r2 + 1
    return blocks

# ---------- Detection: summary + assignments tables ----------
def detect_expected_tables(input_xlsx: str) -> List[Dict[str, Any]]:
    wb = load_workbook(input_xlsx, data_only=True)
    results: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        max_r, max_c = ws.max_row, ws.max_column
        r = 1
        while r <= max_r:
            headers = [ws.cell(r, c).value for c in range(1, max_c + 1)]
            if not any(headers):
                r += 1
                continue

            matched_summary, missing_summary = _match_header_groups(headers, SUMMARY_HEADER_GROUPS)
            matched_assign, missing_assign = _match_header_groups(headers, ASSIGN_HEADER_GROUPS)
            date_cols = sum(1 for h in headers if _is_week_header(h))

            summary_ok = (
                "activity id" in matched_summary and
                ("finish" in matched_summary or "bl project finish" in matched_summary) and
                len(matched_summary) >= 3
            )
            assign_ok = (
                {"activity id", "activity name", "start", "finish"}.issubset(set(matched_assign)) and
                len(matched_assign) >= 5 and
                date_cols >= 5
            )

            if not summary_ok and not assign_ok:
                r += 1
                continue

            nz = [i + 1 for i, v in enumerate(headers) if v not in (None, "", " ")]
            if not nz:
                r += 1
                continue
            c1, c2 = min(nz), max(nz)

            r2 = r + 1
            while r2 <= max_r:
                row_vals = [ws.cell(r2, c).value for c in range(c1, c2 + 1)]
                if all(v in (None, "", " ") for v in row_vals):
                    break
                r2 += 1

            results.append({
                "sheet": ws.title,
                "range": f"R{r}C{c1}:R{r2-1}C{c2}",
                "header_row": r,
                "type": "activity_summary" if summary_ok else "resource_assignments",
                "missing": missing_summary if summary_ok else missing_assign,
                "date_columns": date_cols,
            })
            r = r2 + 1
    return results

def _parse_range(range_str: str) -> Tuple[int, int, int, int]:
    m = re.match(r"R(\d+)C(\d+):R(\d+)C(\d+)", range_str)
    if not m:
        raise ValueError(f"Invalid range: {range_str}")
    return tuple(int(x) for x in m.groups())

def compare_activity_ids(
    input_xlsx: str,
    column_mapping: dict[str, dict[str, str]] | None = None,
) -> Dict[str, Any]:
    wb = load_workbook(input_xlsx, data_only=True)
    tables = detect_expected_tables(input_xlsx)
    summary_ids: List[str] = []
    assign_ids: List[str] = []

    for t in tables:
        ws = wb[t["sheet"]]
        r1, c1, r2, c2 = _parse_range(t["range"])
        header = [ws.cell(r1, c).value for c in range(c1, c2 + 1)]
        mapping = (column_mapping or {}).get(t["type"])
        if mapping:
            _, header = _apply_column_mapping(pd.DataFrame(), header, mapping)
        id_idx = None
        for idx, v in enumerate(header):
            if str(v).strip() == "Activity ID":
                id_idx = idx
                break
        if id_idx is None:
            continue
        for r in range(r1 + 1, r2 + 1):
            val = ws.cell(r, c1 + id_idx).value
            if val is None or str(val).strip() == "":
                continue
            if t["type"] == "activity_summary":
                summary_ids.append(str(val))
            elif t["type"] == "resource_assignments":
                assign_ids.append(str(val))

    summary_set = {s.strip() for s in summary_ids}
    assign_set = {s.strip() for s in assign_ids}
    return {
        "summary_unique": len(summary_set),
        "assign_unique": len(assign_set),
        "summary_only": sorted(summary_set - assign_set),
        "assign_only": sorted(assign_set - summary_set),
    }

def _row_count(t: Dict[str, Any]) -> int:
    r1, _, r2, _ = _parse_range(t["range"])
    return r2 - r1

def _load_table_from_meta(
    wb: Any, table: Dict[str, Any]
) -> Tuple[pd.DataFrame, Dict[str, Any], list[Any]]:
    ws = wb[table["sheet"]]
    r1, c1, r2, c2 = _parse_range(table["range"])
    raw_headers = [ws.cell(r1, c).value for c in range(c1, c2 + 1)]
    data = [[ws.cell(rr, cc).value for cc in range(c1, c2 + 1)] for rr in range(r1, r2 + 1)]
    df = pd.DataFrame(data)
    df.columns = make_unique_columns([str(x).strip() if x is not None else "" for x in raw_headers])
    df = df.iloc[1:, :].reset_index(drop=True)
    df, top_trim, left_trim, _, right_trim = trim_empty_border_with_offsets(df)
    if right_trim:
        raw_headers = raw_headers[left_trim:-right_trim]
    else:
        raw_headers = raw_headers[left_trim:]
    meta = {
        "sheet": table["sheet"],
        "range": table["range"],
        "header_row": r1,
        "data_row_start": r1 + 1 + top_trim,
        "data_col_start": c1 + left_trim,
    }
    return df, meta, raw_headers

def _load_detected_table(
    input_xlsx: str,
    table_type: str,
    column_mapping: dict[str, dict[str, str]] | None = None,
) -> Tuple[pd.DataFrame, Dict[str, Any], list[Any]] | None:
    wb = load_workbook(input_xlsx, data_only=True)
    tables = [t for t in detect_expected_tables(input_xlsx) if t["type"] == table_type]
    if not tables:
        return None
    table = max(tables, key=_row_count)
    df, meta, raw_headers = _load_table_from_meta(wb, table)
    mapping = (column_mapping or {}).get(table_type)
    df, raw_headers = _apply_column_mapping(df, raw_headers, mapping)
    return df, meta, raw_headers

def _load_resource_assignments_table(
    input_xlsx: str,
    spreadsheet_field_marker: str | None,
    column_mapping: dict[str, dict[str, str]] | None = None,
) -> Tuple[pd.DataFrame, Dict[str, Any], list[Any], bool] | None:
    wb = load_workbook(input_xlsx, data_only=True)
    tables = [t for t in detect_expected_tables(input_xlsx) if t["type"] == "resource_assignments"]
    if not tables:
        return None

    matches: List[Dict[str, Any]] = []
    if spreadsheet_field_marker:
        marker = spreadsheet_field_marker.lower()
        for t in tables:
            ws = wb[t["sheet"]]
            r1, c1, r2, c2 = _parse_range(t["range"])
            raw_headers = [ws.cell(r1, c).value for c in range(c1, c2 + 1)]
            mapping = (column_mapping or {}).get("resource_assignments")
            if mapping:
                _, raw_headers = _apply_column_mapping(pd.DataFrame(), raw_headers, mapping)
            field_idx = _find_header_idx(raw_headers, "Spreadsheet Field")
            if field_idx is None:
                continue
            for r in range(r1 + 1, r2 + 1):
                val = ws.cell(r, c1 + field_idx).value
                if val is None:
                    continue
                if marker in str(val).lower():
                    matches.append(t)
                    break

    candidates = matches if matches else tables
    table = max(candidates, key=_row_count)
    df, meta, raw_headers = _load_table_from_meta(wb, table)
    mapping = (column_mapping or {}).get("resource_assignments")
    df, raw_headers = _apply_column_mapping(df, raw_headers, mapping)
    matched = bool(matches) and table in matches
    meta["marker"] = spreadsheet_field_marker
    meta["marker_matched"] = matched
    return df, meta, raw_headers, matched

def build_schedule_lookup(
    input_xlsx: str,
    today: date | None = None,
    column_mapping: dict[str, dict[str, str]] | None = None,
) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Any]]:
    """
    Calcule Schedule % depuis le tableau Ressource Assignments :
      Schedule % = valeur de la semaine courante / Budgeted Units * 100
    """
    info: Dict[str, Any] = {
        "status": "ok",
        "week_date": None,
        "week_col": None,
        "table": None,
        "errors": [],
    }
    table = _load_resource_assignments_table(
        input_xlsx,
        "Cum Budgeted Units",
        column_mapping=column_mapping,
    )
    if table is None:
        info["status"] = "missing_table"
        info["errors"].append("Resource assignments table not found.")
        return {}, info

    df, meta, raw_headers, matched = table
    info["table"] = meta
    if not matched:
        info["errors"].append("Planned table not found by Spreadsheet Field = Cum Budgeted Units.")

    id_idx = _find_header_idx(raw_headers, "Activity ID")
    budget_idx = _find_header_idx(raw_headers, "Budgeted Units")
    if id_idx is None or budget_idx is None:
        info["status"] = "missing_columns"
        info["errors"].append("Missing Activity ID or Budgeted Units columns in resource assignments.")
        return {}, info

    today = today or date.today()
    target_week = _week_start(today)
    info["week_date"] = target_week.isoformat()
    week_idx = None
    for idx, h in enumerate(raw_headers):
        h_date = _to_excel_date(h)
        if h_date and h_date == target_week:
            week_idx = idx
            break
    if week_idx is None:
        info["status"] = "week_not_found"
        info["errors"].append(f"No column for current week ({target_week.isoformat()}).")
    else:
        info["week_col"] = str(raw_headers[week_idx])

    id_col = df.columns[id_idx]
    budget_col = df.columns[budget_idx]
    week_col = df.columns[week_idx] if week_idx is not None else None

    lookup: Dict[str, Dict[str, Any]] = {}
    for row_idx, row in df.iterrows():
        raw_id = row.get(id_col)
        if raw_id is None or str(raw_id).strip() == "":
            continue
        key = str(raw_id).strip()
        if key in lookup:
            continue
        budget = _safe_float(row.get(budget_col))
        week_val = _safe_float(row.get(week_col)) if week_col else None
        budget_cell = _cell_ref(meta, row_idx, budget_idx) if budget_idx is not None else None
        week_cell = _cell_ref(meta, row_idx, week_idx) if week_idx is not None else None

        tip = "Schedule % = Units (current week) / Budgeted Units"
        display = "?"
        value = None
        if week_idx is None:
            tip = f"Schedule unavailable: week column {target_week.isoformat()} not found."
        elif budget in (None, 0):
            tip = "Schedule unavailable: Budgeted Units missing or 0."
        elif week_val is None:
            tip = "Schedule unavailable: week cell is empty."
        else:
            value = (week_val / budget) * 100.0
            display = f"{value:.2f}%"
            tip = f"Schedule % = Units ({target_week.isoformat()}) / Budgeted Units"
        tip = _append_tip_sources(
            tip,
            [
                f"Week: {week_cell}" if week_cell else "",
                f"Budgeted Units: {budget_cell}" if budget_cell else "",
            ],
        )

        lookup[key] = {
            "value": value,
            "display": display,
            "tip": tip,
            "budgeted_units": budget,
            "budget_cell": budget_cell,
            "week_cell": week_cell,
        }

    return lookup, info

def build_weekly_progress(
    input_xlsx: str,
    activity_id: str,
    today: date | None = None,
    column_mapping: dict[str, dict[str, str]] | None = None,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Weekly planned progress per activity:
      Planned % = (week_value - previous_week_value) / Budgeted Units * 100
    """
    info: Dict[str, Any] = {
        "status": "ok",
        "week_date": None,
        "current_week_date": None,
        "current_week_label": None,
        "table": None,
        "errors": [],
    }
    planned_table = _load_resource_assignments_table(
        input_xlsx,
        "Cum Budgeted Units",
        column_mapping=column_mapping,
    )
    if planned_table is None:
        info["status"] = "missing_table"
        info["errors"].append("Resource assignments table not found.")
        return [], info

    df, meta, raw_headers, planned_matched = planned_table
    info["table"] = meta
    if not planned_matched:
        info["errors"].append("Planned table not found by Spreadsheet Field = Cum Budgeted Units.")

    actual_past_table = _load_resource_assignments_table(
        input_xlsx,
        "Cum Actual Units",
        column_mapping=column_mapping,
    )
    actual_future_table = _load_resource_assignments_table(
        input_xlsx,
        "Cum Remaining Early Units",
        column_mapping=column_mapping,
    )
    actual_missing_reason_past = None
    actual_missing_reason_future = None

    actual_past_df = None
    actual_past_headers: list[Any] = []
    actual_past_meta: Dict[str, Any] | None = None
    actual_past_matched = False

    actual_future_df = None
    actual_future_headers: list[Any] = []
    actual_future_meta: Dict[str, Any] | None = None
    actual_future_matched = False

    if actual_past_table is None:
        info["errors"].append("Actual table not found by Spreadsheet Field = Cum Actual Units.")
        actual_missing_reason_past = "Actual unavailable: table with Cum Actual Units not found."
    else:
        actual_past_df, actual_past_meta, actual_past_headers, actual_past_matched = actual_past_table
        info["actual_table_past"] = actual_past_meta
        if not actual_past_matched:
            info["errors"].append("Actual table not found by Spreadsheet Field = Cum Actual Units.")
            actual_missing_reason_past = "Actual unavailable: table with Cum Actual Units not found."
            actual_past_df = None

    if actual_future_table is None:
        info["errors"].append("Actual table not found by Spreadsheet Field = Cum Remaining Early Units.")
        actual_missing_reason_future = "Actual unavailable: table with Cum Remaining Early Units not found."
    else:
        actual_future_df, actual_future_meta, actual_future_headers, actual_future_matched = actual_future_table
        info["actual_table_future"] = actual_future_meta
        if not actual_future_matched:
            info["errors"].append("Actual table not found by Spreadsheet Field = Cum Remaining Early Units.")
            actual_missing_reason_future = "Actual unavailable: table with Cum Remaining Early Units not found."
            actual_future_df = None

    id_idx = _find_header_idx(raw_headers, "Activity ID")
    budget_idx = _find_header_idx(raw_headers, "Budgeted Units")
    if id_idx is None or budget_idx is None:
        info["status"] = "missing_columns"
        info["errors"].append("Missing Activity ID or Budgeted Units columns in resource assignments.")
        return [], info

    planned_week_map: Dict[date, int] = {}
    planned_label_map: Dict[date, Any] = {}
    for idx, h in enumerate(raw_headers):
        h_date = _to_excel_date(h)
        if not h_date:
            continue
        week = _week_start(h_date)
        if week not in planned_week_map:
            planned_week_map[week] = idx
            planned_label_map[week] = h

    today = today or date.today()
    target_week = _week_start(today)
    info["week_date"] = target_week.isoformat()
    info["current_week_date"] = target_week.isoformat()

    id_col = df.columns[id_idx]
    budget_col = df.columns[budget_idx]
    actual_past_id_col = None
    actual_past_week_map: Dict[date, int] = {}
    actual_past_label_map: Dict[date, Any] = {}
    actual_past_row = None
    actual_past_row_idx = None
    actual_past_missing = actual_past_df is None

    if actual_past_df is not None:
        actual_past_id_idx = _find_header_idx(actual_past_headers, "Activity ID")
        if actual_past_id_idx is None:
            actual_past_missing = True
            info["errors"].append("Missing Activity ID column in Cum Actual Units table.")
            actual_missing_reason_past = "Actual unavailable: missing Activity ID (Cum Actual Units)."
        else:
            actual_past_id_col = actual_past_df.columns[actual_past_id_idx]
            for idx, h in enumerate(actual_past_headers):
                h_date = _to_excel_date(h)
                if not h_date:
                    continue
                week = _week_start(h_date)
                if week not in actual_past_week_map:
                    actual_past_week_map[week] = idx
                    actual_past_label_map[week] = h
            for idx, r in actual_past_df.iterrows():
                raw_id = r.get(actual_past_id_col)
                if raw_id is None or str(raw_id).strip() == "":
                    continue
                if str(raw_id).strip() == str(activity_id).strip():
                    actual_past_row = r
                    actual_past_row_idx = idx
                    break
            if actual_past_row is None:
                actual_past_missing = True
                info["errors"].append("Activity ID not found in Cum Actual Units table.")
                actual_missing_reason_past = "Actual unavailable: Activity ID not found in Cum Actual Units table."
            else:
                pass
    actual_future_id_col = None
    actual_future_week_map: Dict[date, int] = {}
    actual_future_label_map: Dict[date, Any] = {}
    actual_future_row = None
    actual_future_row_idx = None
    actual_future_missing = actual_future_df is None

    if actual_future_df is not None:
        actual_future_id_idx = _find_header_idx(actual_future_headers, "Activity ID")
        if actual_future_id_idx is None:
            actual_future_missing = True
            info["errors"].append("Missing Activity ID column in Cum Remaining Early Units table.")
            actual_missing_reason_future = "Actual unavailable: missing Activity ID (Cum Remaining Early Units)."
        else:
            actual_future_id_col = actual_future_df.columns[actual_future_id_idx]
            for idx, h in enumerate(actual_future_headers):
                h_date = _to_excel_date(h)
                if not h_date:
                    continue
                week = _week_start(h_date)
                if week not in actual_future_week_map:
                    actual_future_week_map[week] = idx
                    actual_future_label_map[week] = h
            for idx, r in actual_future_df.iterrows():
                raw_id = r.get(actual_future_id_col)
                if raw_id is None or str(raw_id).strip() == "":
                    continue
                if str(raw_id).strip() == str(activity_id).strip():
                    actual_future_row = r
                    actual_future_row_idx = idx
                    break
            if actual_future_row is None:
                actual_future_missing = True
                info["errors"].append("Activity ID not found in Cum Remaining Early Units table.")
                actual_missing_reason_future = "Actual unavailable: Activity ID not found in Cum Remaining Early Units table."
            else:
                pass

    row = None
    row_idx = None
    for idx, r in df.iterrows():
        raw_id = r.get(id_col)
        if raw_id is None or str(raw_id).strip() == "":
            continue
        if str(raw_id).strip() == str(activity_id).strip():
            row = r
            row_idx = idx
            break

    activity_missing = row is None
    if activity_missing:
        info["status"] = "activity_not_found"
        info["errors"].append("Activity ID not found in resource assignments.")

    budget = _safe_float(row.get(budget_col)) if row is not None else None
    budget_cell = _cell_ref(meta, row_idx, budget_idx) if row_idx is not None else None
    if budget in (None, 0):
        if not activity_missing:
            info["errors"].append("Budgeted Units missing or 0 for selected activity.")

    week_dates = sorted(
        set(planned_week_map.keys())
        | set(actual_past_week_map.keys())
        | set(actual_future_week_map.keys())
    )
    if not week_dates:
        info["status"] = "missing_week_columns"
        info["errors"].append("No weekly date columns found in resource assignments.")
        return [], info

    planned_start_week = min(planned_week_map.keys()) if planned_week_map else None
    planned_end_week = max(planned_week_map.keys()) if planned_week_map else None
    if planned_start_week:
        info["planned_start_week"] = planned_start_week.isoformat()
    if planned_end_week:
        info["planned_end_week"] = planned_end_week.isoformat()
    info["available_end_week"] = week_dates[-1].isoformat() if week_dates else None
    info["table_ranges"] = {
        "planned": info.get("table"),
        "actual_past": info.get("actual_table_past"),
        "actual_future": info.get("actual_table_future"),
    }

    if target_week not in week_dates:
        info["errors"].append(f"Current week {target_week.isoformat()} not found in weekly columns.")
    else:
        weeks_before = len([w for w in week_dates if w < target_week])
        weeks_after = len([w for w in week_dates if w > target_week])
        info["weeks_before"] = weeks_before
        info["weeks_after"] = weeks_after
        if weeks_before < 3 or weeks_after < 3:
            info["errors"].append(
                "Current week cannot be centered "
                f"(need 3 weeks before/after, have {weeks_before} before, {weeks_after} after; "
                f"available range {info.get('planned_start_week')} to {info.get('available_end_week')}; "
                f"tables planned={info.get('table')}, "
                f"actual_past={info.get('actual_table_past')}, "
                f"actual_future={info.get('actual_table_future')})."
            )

    series: List[Dict[str, Any]] = []
    prev_val = 0.0
    prev_actual_val_past = 0.0
    prev_actual_val_future = 0.0
    for week_date in week_dates:
        header = (
            planned_label_map.get(week_date)
            or actual_past_label_map.get(week_date)
            or actual_future_label_map.get(week_date)
        )
        week_label = str(header).strip() if header not in (None, "") else week_date.strftime("%d-%b-%y")
        if week_date == target_week:
            info["current_week_label"] = week_label
        planned_week_idx = planned_week_map.get(week_date)
        week_val = (
            _safe_float(row.get(df.columns[planned_week_idx]))
            if row is not None and planned_week_idx is not None
            else None
        )
        planned_week_cell = (
            _cell_ref(meta, row_idx, planned_week_idx)
            if row_idx is not None and planned_week_idx is not None
            else None
        )
        actual_past_week_val = None
        actual_past_week_idx = None
        actual_past_week_cell = None
        if not actual_past_missing and week_date in actual_past_week_map:
            actual_past_week_idx = actual_past_week_map[week_date]
            actual_past_week_val = _safe_float(actual_past_row.get(actual_past_df.columns[actual_past_week_idx])) if actual_past_row is not None else None
            actual_past_week_cell = (
                _cell_ref(actual_past_meta, actual_past_row_idx, actual_past_week_idx)
                if actual_past_row_idx is not None and actual_past_meta is not None
                else None
            )

        actual_future_week_val = None
        actual_future_week_idx = None
        actual_future_week_cell = None
        if not actual_future_missing and week_date in actual_future_week_map:
            actual_future_week_idx = actual_future_week_map[week_date]
            actual_future_week_val = _safe_float(actual_future_row.get(actual_future_df.columns[actual_future_week_idx])) if actual_future_row is not None else None
            actual_future_week_cell = (
                _cell_ref(actual_future_meta, actual_future_row_idx, actual_future_week_idx)
                if actual_future_row_idx is not None and actual_future_meta is not None
                else None
            )

        planned_val = None
        planned_display = "?"
        planned_tip = "Planned % = (Units this week - Units previous week) / Budgeted Units"
        planned_cum_val = None
        planned_cum_display = "?"
        planned_cum_tip = "Cumulative Planned % = Units (week) / Budgeted Units"

        if activity_missing:
            planned_tip = "Planned unavailable: Activity ID not found in resource assignments."
            planned_cum_tip = "Cumulative planned unavailable: Activity ID not found in resource assignments."
        elif planned_week_idx is None:
            if planned_end_week and week_date > planned_end_week:
                planned_val = 0.0
                planned_display = "0.00%"
                planned_tip = "Planned % = 0 (outside baseline date range)."
                planned_cum_tip = "Cumulative planned unavailable: week is outside baseline date range."
            else:
                planned_tip = f"Planned unavailable: week column {week_date.isoformat()} not found."
                planned_cum_tip = f"Cumulative planned unavailable: week column {week_date.isoformat()} not found."
        elif budget in (None, 0):
            planned_tip = "Planned unavailable: Budgeted Units missing or 0."
            planned_cum_tip = "Cumulative planned unavailable: Budgeted Units missing or 0."
        elif week_val is None:
            planned_tip = "Planned unavailable: week cell is empty."
            planned_cum_tip = "Cumulative planned unavailable: week cell is empty."
        else:
            delta = week_val - prev_val
            if delta < 0:
                planned_tip = "Planned unavailable: week value decreased vs previous week."
            else:
                planned_val = (delta / budget) * 100.0
                planned_display = f"{planned_val:.2f}%"
            planned_cum_val = (week_val / budget) * 100.0
            planned_cum_display = f"{planned_cum_val:.2f}%"
        planned_sources = [
            f"Week: {planned_week_cell}" if planned_week_cell else "",
            f"Budgeted Units: {budget_cell}" if budget_cell else "",
        ]
        planned_tip = _append_tip_sources(planned_tip, planned_sources)
        planned_cum_tip = _append_tip_sources(planned_cum_tip, planned_sources)
        if week_val is not None:
            prev_val = week_val

        use_future = week_date >= target_week
        actual_val = None
        actual_display = "?"
        actual_cum_val = None
        actual_cum_display = "?"
        actual_cum_actual_val = None
        actual_cum_actual_display = "?"
        actual_cum_actual_tip = "Cumulative Actual % = Units (week) / Budgeted Units"
        if use_future:
            actual_tip = "Actual % = (Units this week - Units previous week) / Budgeted Units"
            actual_cum_tip = "Cumulative Actual % = Units (week) / Budgeted Units"
            if actual_future_missing:
                actual_tip = actual_missing_reason_future or "Actual unavailable: table with Cum Remaining Early Units not found."
                actual_cum_tip = actual_tip.replace("Actual %", "Cumulative Actual %")
            elif actual_future_week_idx is None:
                actual_tip = f"Actual unavailable: week column {week_date.isoformat()} not found (Cum Remaining Early Units)."
                actual_cum_tip = f"Cumulative actual unavailable: week column {week_date.isoformat()} not found (Cum Remaining Early Units)."
            elif budget in (None, 0):
                actual_tip = "Actual unavailable: Budgeted Units missing or 0."
                actual_cum_tip = "Cumulative actual unavailable: Budgeted Units missing or 0."
            elif actual_future_week_val is None:
                actual_tip = "Actual unavailable: week cell is empty (Cum Remaining Early Units)."
                actual_cum_tip = "Cumulative actual unavailable: week cell is empty (Cum Remaining Early Units)."
            else:
                delta_actual = actual_future_week_val - prev_actual_val_future
                if delta_actual < 0:
                    actual_tip = "Actual unavailable: week value decreased vs previous week (Cum Remaining Early Units)."
                    actual_cum_tip = "Cumulative actual unavailable: week value decreased vs previous week (Cum Remaining Early Units)."
                else:
                    actual_val = (delta_actual / budget) * 100.0
                    actual_display = f"{actual_val:.2f}%"
                actual_cum_val = (actual_future_week_val / budget) * 100.0
                actual_cum_display = f"{actual_cum_val:.2f}%"
            actual_sources = [
                f"Week: {actual_future_week_cell}" if actual_future_week_cell else "",
                f"Budgeted Units: {budget_cell}" if budget_cell else "",
            ]
        else:
            actual_tip = "Actual % = (Units this week - Units previous week) / Budgeted Units"
            actual_cum_tip = "Cumulative Actual % = Units (week) / Budgeted Units"
            if actual_past_missing:
                actual_tip = actual_missing_reason_past or "Actual unavailable: table with Cum Actual Units not found."
                actual_cum_tip = actual_tip.replace("Actual %", "Cumulative Actual %")
            elif actual_past_week_idx is None:
                actual_tip = f"Actual unavailable: week column {week_date.isoformat()} not found (Cum Actual Units)."
                actual_cum_tip = f"Cumulative actual unavailable: week column {week_date.isoformat()} not found (Cum Actual Units)."
            elif budget in (None, 0):
                actual_tip = "Actual unavailable: Budgeted Units missing or 0."
                actual_cum_tip = "Cumulative actual unavailable: Budgeted Units missing or 0."
            elif actual_past_week_val is None:
                actual_tip = "Actual unavailable: week cell is empty (Cum Actual Units)."
                actual_cum_tip = "Cumulative actual unavailable: week cell is empty (Cum Actual Units)."
            else:
                delta_actual = actual_past_week_val - prev_actual_val_past
                if delta_actual < 0:
                    actual_tip = "Actual unavailable: week value decreased vs previous week (Cum Actual Units)."
                    actual_cum_tip = "Cumulative actual unavailable: week value decreased vs previous week (Cum Actual Units)."
                else:
                    actual_val = (delta_actual / budget) * 100.0
                    actual_display = f"{actual_val:.2f}%"
                actual_cum_val = (actual_past_week_val / budget) * 100.0
                actual_cum_display = f"{actual_cum_val:.2f}%"
            actual_sources = [
                f"Week: {actual_past_week_cell}" if actual_past_week_cell else "",
                f"Budgeted Units: {budget_cell}" if budget_cell else "",
            ]
        actual_tip = _append_tip_sources(actual_tip, actual_sources)
        actual_cum_tip = _append_tip_sources(actual_cum_tip, actual_sources)
        actual_cum_actual_sources = [
            f"Week: {actual_past_week_cell}" if actual_past_week_cell else "",
            f"Budgeted Units: {budget_cell}" if budget_cell else "",
        ]
        if actual_past_missing:
            actual_cum_actual_tip = actual_missing_reason_past or "Cumulative actual unavailable: table with Cum Actual Units not found."
        elif actual_past_week_idx is None:
            actual_cum_actual_tip = f"Cumulative actual unavailable: week column {week_date.isoformat()} not found (Cum Actual Units)."
        elif budget in (None, 0):
            actual_cum_actual_tip = "Cumulative actual unavailable: Budgeted Units missing or 0."
        elif actual_past_week_val is None:
            actual_cum_actual_tip = "Cumulative actual unavailable: week cell is empty (Cum Actual Units)."
        else:
            actual_cum_actual_val = (actual_past_week_val / budget) * 100.0
            actual_cum_actual_display = f"{actual_cum_actual_val:.2f}%"
        actual_cum_actual_tip = _append_tip_sources(actual_cum_actual_tip, actual_cum_actual_sources)

        if actual_past_week_val is not None:
            prev_actual_val_past = actual_past_week_val
        if actual_future_week_val is not None:
            prev_actual_val_future = actual_future_week_val

        if week_date == target_week and actual_display == "?":
            current_source = "Cum Remaining Early Units" if use_future else "Cum Actual Units"
            detail = actual_tip or "Actual unavailable for current week."
            warn = f"Actual current week unavailable ({current_source}). {detail}"
            if warn not in info["errors"]:
                info["errors"].append(warn)

        series.append(
            {
                "week": week_label,
                "week_date": week_date,
                "week_label": week_date.strftime("%d-%b"),
                "planned": planned_val,
                "planned_display": planned_display,
                "planned_tip": planned_tip,
                "planned_cum": planned_cum_val,
                "planned_cum_display": planned_cum_display,
                "planned_cum_tip": planned_cum_tip,
                "actual": actual_val,
                "actual_display": actual_display,
                "actual_tip": actual_tip,
                "actual_cum": actual_cum_val,
                "actual_cum_display": actual_cum_display,
                "actual_cum_tip": actual_cum_tip,
                "actual_cum_actual": actual_cum_actual_val,
                "actual_cum_actual_display": actual_cum_actual_display,
                "actual_cum_actual_tip": actual_cum_actual_tip,
                "actual_cum_units": actual_past_week_val if not use_future else None,
                "actual_week_cell": actual_past_week_cell if not use_future else None,
                "budgeted_units": budget,
                "budgeted_units_cell": budget_cell,
                "forecast_cum_units": actual_future_week_val if use_future else None,
                "forecast_week_cell": actual_future_week_cell if use_future else None,
            }
        )

    return series, info

def _find_header_idx_norm(headers: list[Any], candidates: list[str]) -> int | None:
    norm_headers = [_norm_header(h) for h in headers]
    for cand in candidates:
        cand_norm = _norm_header(cand)
        for idx, h in enumerate(norm_headers):
            if h == cand_norm:
                return idx
    return None

def build_preview_rows(
    input_xlsx: str,
    table_type: str = "activity_summary",
    prefer_first_table: bool = False,
    column_mapping: dict[str, dict[str, str]] | None = None,
) -> List[Dict[str, Any]]:
    wb = load_workbook(input_xlsx, data_only=True)
    tables = [t for t in detect_expected_tables(input_xlsx) if t["type"] == table_type]
    rows: List[Dict[str, Any]] = []
    mapping = (column_mapping or {}).get(table_type, {})
    field_variants = _table_field_variants(table_type)

    def _lead_spaces(s: str) -> int:
        return len(re.match(r"^\s*", s).group(0))

    def _idx(headers: list[Any], canonical: str) -> int | None:
        mapped = mapping.get(canonical)
        if mapped:
            return _find_header_idx_norm(headers, [mapped])
        variants = field_variants.get(canonical, [canonical])
        return _find_header_idx_norm(headers, variants)

    if prefer_first_table:
        for ws in wb.worksheets:
            max_r, max_c = ws.max_row, ws.max_column
            for r in range(1, max_r + 1):
                headers = [ws.cell(r, c).value for c in range(1, max_c + 1)]
                if not any(headers):
                    continue
                nz = [i + 1 for i, v in enumerate(headers) if v not in (None, "", " ")]
                if not nz:
                    continue
                c1, c2 = min(nz), max(nz)
                header = [ws.cell(r, c).value for c in range(c1, c2 + 1)]
                id_idx = _idx(header, "Activity ID")
                if id_idx is None:
                    continue
                name_idx = _idx(header, "Activity Name")
                status_idx = _idx(header, "Activity Status")
                units_idx = _idx(header, "Units % Complete")
                bl_finish_idx = _idx(header, "BL Project Finish")
                finish_idx = _idx(header, "Finish")
                variance_idx = _idx(header, "Variance - BL Project Finish Date")
                budget_idx = _idx(header, "Budgeted Labor Units")

                r2 = r + 1
                while r2 <= max_r:
                    row_vals = [ws.cell(r2, c).value for c in range(c1, c2 + 1)]
                    if all(v in (None, "", " ") for v in row_vals):
                        break
                    r2 += 1

                rows = []
                sheet_ref = _sheet_ref(ws.title)
                for rr in range(r + 1, r2):
                    id_val = ws.cell(rr, c1 + id_idx).value
                    if id_val is None or str(id_val).strip() == "":
                        continue
                    raw_id = str(id_val)
                    activity_id = raw_id.strip()
                    name_val = ws.cell(rr, c1 + name_idx).value if name_idx is not None else None
                    name_text = str(name_val).strip() if name_val is not None else ""
                    display_label = f"{activity_id} - {name_text}".strip(" -")
                    rows.append(
                        {
                            "sheet": ws.title,
                            "range": f"R{r}C{c1}:R{r2-1}C{c2}",
                            "raw": raw_id,
                            "label": name_text if name_text else activity_id,
                            "display_label": display_label,
                            "indent": _lead_spaces(raw_id),
                            "activity_name": name_text,
                            "activity_id": activity_id,
                            "activity_status": ws.cell(rr, c1 + status_idx).value if status_idx is not None else None,
                            "units_complete": ws.cell(rr, c1 + units_idx).value if units_idx is not None else None,
                            "bl_project_finish": ws.cell(rr, c1 + bl_finish_idx).value if bl_finish_idx is not None else None,
                            "finish": ws.cell(rr, c1 + finish_idx).value if finish_idx is not None else None,
                            "variance_days": ws.cell(rr, c1 + variance_idx).value if variance_idx is not None else None,
                            "budgeted_units": ws.cell(rr, c1 + budget_idx).value if budget_idx is not None else None,
                            "units_complete_cell": f"{sheet_ref}!{get_column_letter(c1 + units_idx)}{rr}" if units_idx is not None else None,
                            "bl_project_finish_cell": f"{sheet_ref}!{get_column_letter(c1 + bl_finish_idx)}{rr}" if bl_finish_idx is not None else None,
                            "finish_cell": f"{sheet_ref}!{get_column_letter(c1 + finish_idx)}{rr}" if finish_idx is not None else None,
                            "variance_days_cell": f"{sheet_ref}!{get_column_letter(c1 + variance_idx)}{rr}" if variance_idx is not None else None,
                        }
                    )

                if rows:
                    indents = sorted({r["indent"] for r in rows})
                    indent_to_level = {sp: i for i, sp in enumerate(indents)}
                    for row in rows:
                        row["level"] = indent_to_level.get(row["indent"], 0)
                    return rows

    if not tables:
        return []

    def _row_count(t: Dict[str, Any]) -> int:
        r1, _, r2, _ = _parse_range(t["range"])
        return r2 - r1

    table = max(tables, key=_row_count)
    ws = wb[table["sheet"]]
    r1, c1, r2, c2 = _parse_range(table["range"])
    header = [ws.cell(r1, c).value for c in range(c1, c2 + 1)]
    id_idx = _idx(header, "Activity ID")
    units_idx = _idx(header, "Units % Complete")
    variance_idx = _idx(header, "Variance - BL Project Finish Date")
    bl_finish_idx = _idx(header, "BL Project Finish")
    finish_idx = _idx(header, "Finish")
    if id_idx is None:
        return []

    for r in range(r1 + 1, r2 + 1):
        val = ws.cell(r, c1 + id_idx).value
        if val is None or str(val).strip() == "":
            continue
        raw = str(val)
        units_val = ws.cell(r, c1 + units_idx).value if units_idx is not None else None
        variance_val = ws.cell(r, c1 + variance_idx).value if variance_idx is not None else None
        bl_finish_val = ws.cell(r, c1 + bl_finish_idx).value if bl_finish_idx is not None else None
        finish_val = ws.cell(r, c1 + finish_idx).value if finish_idx is not None else None
        sheet_ref = _sheet_ref(table["sheet"])
        units_cell = f"{sheet_ref}!{get_column_letter(c1 + units_idx)}{r}" if units_idx is not None else None
        variance_cell = f"{sheet_ref}!{get_column_letter(c1 + variance_idx)}{r}" if variance_idx is not None else None
        bl_finish_cell = f"{sheet_ref}!{get_column_letter(c1 + bl_finish_idx)}{r}" if bl_finish_idx is not None else None
        finish_cell = f"{sheet_ref}!{get_column_letter(c1 + finish_idx)}{r}" if finish_idx is not None else None
        rows.append({
            "sheet": table["sheet"],
            "range": table["range"],
            "raw": raw,
            "label": raw.strip(),
            "indent": _lead_spaces(raw),
            "activity_id": raw.strip(),
            "units_complete": units_val,
            "variance_days": variance_val,
            "bl_project_finish": bl_finish_val,
            "finish": finish_val,
            "units_complete_cell": units_cell,
            "variance_days_cell": variance_cell,
            "bl_project_finish_cell": bl_finish_cell,
            "finish_cell": finish_cell,
        })

    if not rows:
        return []

    indents = sorted({r["indent"] for r in rows})
    indent_to_level = {sp: i for i, sp in enumerate(indents)}
    for r in rows:
        r["level"] = indent_to_level.get(r["indent"], 0)
    return rows

# ---------- Choix de la colonne Label ----------
def pick_label_col(df: pd.DataFrame) -> str:
    """
    Priorité :
      1) Colonne d'en-tête vide "" si surtout du texte non date-like
      2) Sinon meilleure colonne non-requise : textuelle, non date-like, variété d'indentation
      3) Sinon première colonne
    """
    req = {"Planned Finish","Forecast Finish","Schedule %","Earned %","ecart","impact","Glissement"}
    if "" in df.columns:
        col = df[""].astype(str)
        if (col.str.strip() != "").any():
            not_dates_ratio = (~col.str.strip().apply(_is_date_like)).mean()
            if not_dates_ratio > 0.6:
                return ""

    best, best_score = None, -1.0
    for c in df.columns:
        if c in req: 
            continue
        col = df[c].astype(str)
        texty = col.apply(lambda s: any(ch.isalpha() for ch in s)).mean()
        not_date = col.apply(lambda s: not _is_date_like(s)).mean()
        indents = col.apply(leading_spaces)
        indent_levels = indents.nunique()
        score = (texty*2.0) + (not_date*1.5) + (min(indent_levels, 4)*0.8)
        if score > best_score:
            best_score, best = score, c
    return best or df.columns[0]

# ---------- WBS builder ----------
def to_wbs_tree(
    df: pd.DataFrame,
    label_col: str,
    schedule_lookup: Dict[str, Dict[str, Any]] | None = None,
    schedule_info: Dict[str, Any] | None = None,
    source_meta: Dict[str, Any] | None = None,
) -> Dict:
    df = df.copy()
    df["_row_idx"] = df.index
    df[label_col] = df[label_col].fillna("")
    df = df[df[label_col].astype(str).str.strip() != ""]
    df["_indent"] = df[label_col].apply(leading_spaces)

    uniq = sorted(df["_indent"].unique().tolist()) or [0]
    space2lvl = {sp: i+1 for i, sp in enumerate(uniq)}
    activity_id_col = "Activity ID" if "Activity ID" in df.columns else label_col
    root_activity_id = str(df.iloc[0][activity_id_col] or "").strip() if not df.empty else ""
    root_budget = None
    root_budget_cell = None
    if schedule_lookup and root_activity_id in schedule_lookup:
        root_entry = schedule_lookup[root_activity_id]
        root_budget = root_entry.get("budgeted_units")
        root_budget_cell = root_entry.get("budget_cell")

    def _cell_for(col_name: str | None, row_idx: int | None) -> str | None:
        if col_name is None or row_idx is None or source_meta is None:
            return None
        if col_name not in df.columns:
            return None
        col_idx = int(df.columns.get_loc(col_name))
        return _cell_ref(source_meta, row_idx, col_idx)

    def row_metrics(r: pd.Series, row_idx: int | None) -> dict:
        # NOTE: ecart/impact -> ENTIER, schedule/earned -> tidy (int si rond)
        planned_source_col = "BL Project Finish"
        planned = r.get(planned_source_col)
        if planned is None or str(planned).strip() == "":
            planned_source_col = "Planned Finish"
            planned = r.get(planned_source_col)

        forecast_source_col = "Finish"
        forecast = r.get(forecast_source_col)
        if forecast is None or str(forecast).strip() == "":
            forecast_source_col = "Forecast Finish"
            forecast = r.get(forecast_source_col)

        earned_source_col = "Units % Complete"
        earned_raw = r.get(earned_source_col)
        if earned_raw is None or str(earned_raw).strip() == "":
            earned_source_col = "Earned %"
            earned_raw = r.get(earned_source_col)

        schedule_val = None
        schedule_display = "?"
        schedule_tip = None
        activity_id = str(r.get(activity_id_col) or "").strip()
        schedule_week_cell = None
        schedule_budget_cell = None
        activity_budget_cell = None
        if schedule_lookup is not None:
            if activity_id and activity_id in schedule_lookup:
                entry = schedule_lookup[activity_id]
                schedule_val = entry.get("value")
                schedule_display = entry.get("display", "?")
                schedule_tip = entry.get("tip")
                activity_budget = entry.get("budgeted_units")
                schedule_week_cell = entry.get("week_cell")
                schedule_budget_cell = entry.get("budget_cell")
                activity_budget_cell = entry.get("budget_cell")
            else:
                if schedule_info and schedule_info.get("status") in ("missing_table", "missing_columns", "week_not_found"):
                    schedule_tip = "Schedule unavailable: " + " ".join(schedule_info.get("errors", []))
                else:
                    schedule_tip = "Schedule unavailable: Activity ID not found in resource assignments."
                activity_budget = None
        else:
            schedule_tip = "Schedule unavailable: resource assignments not loaded."
            activity_budget = None

        if earned_raw is None or str(earned_raw).strip() == "":
            earned_val = None
            earned_display = "?"
            earned_tip = f"Earned unavailable: {earned_source_col} missing."
        else:
            earned_val = tidy_num(parse_percent_float(earned_raw))
            earned_display = f"{earned_val:.2f}%"
            earned_tip = f"Earned % = {earned_source_col}"

        ecart_val = None
        if isinstance(earned_val, (int, float)) and isinstance(schedule_val, (int, float)):
            ecart_val = earned_val - schedule_val
            ecart_display = f"{ecart_val:+.2f}%"
            ecart_tip = "Variance = Earned % - Schedule %"
        else:
            ecart_display = "?"
            if earned_val is None:
                ecart_tip = "Variance unavailable: Earned % missing."
            elif schedule_val is None:
                ecart_tip = "Variance unavailable: Schedule % missing."
            else:
                ecart_tip = "Variance unavailable: missing values."
        impact_val = None
        if isinstance(ecart_val, (int, float)):
            if root_budget in (None, 0):
                impact_display = "?"
                impact_tip = "Impact unavailable: root Budgeted Units missing or 0."
            elif activity_budget in (None, 0):
                impact_display = "?"
                impact_tip = "Impact unavailable: Budgeted Units missing or 0."
            else:
                impact_val = (activity_budget / root_budget) * ecart_val
                impact_display = f"{impact_val:+.2f}%"
                impact_tip = "Impact = (Budgeted Units / Root Budgeted Units) * Variance"
        else:
            impact_display = "?"
            if ecart_val is None:
                impact_tip = "Impact unavailable: Variance missing."
            else:
                impact_tip = "Impact unavailable: missing values."
        gliss_source_col = "Variance - BL Project Finish Date"
        gliss_raw = r.get(gliss_source_col)
        if gliss_raw is None or str(gliss_raw).strip() == "":
            gliss_source_col = "Glissement"
            gliss_raw = r.get(gliss_source_col)
        gliss_val = _parse_days(gliss_raw)
        if gliss_val is None:
            gliss_display = "?"
            gliss_tip = "Slip unavailable: Variance - BL Project Finish Date missing."
        else:
            gliss_display = f"{int(gliss_val)}d"
            gliss_tip = "Slip = Variance - BL Project Finish Date (days)"
        planned_text = as_text(planned)
        if planned_text:
            planned_display = planned_text
            planned_tip = f"Planned = {planned_source_col}"
        else:
            planned_display = "?"
            planned_tip = f"Planned unavailable: {planned_source_col} missing."
        forecast_text = as_text(forecast)
        if forecast_text:
            forecast_display = forecast_text
            forecast_tip = f"Forecast = {forecast_source_col}"
        else:
            forecast_display = "?"
            forecast_tip = f"Forecast unavailable: {forecast_source_col} missing."

        planned_tip = _append_tip_sources(
            planned_tip,
            [f"{planned_source_col}: {_cell_for(planned_source_col, row_idx)}"] if planned_source_col else [],
        )
        forecast_tip = _append_tip_sources(
            forecast_tip,
            [f"{forecast_source_col}: {_cell_for(forecast_source_col, row_idx)}"] if forecast_source_col else [],
        )
        earned_tip = _append_tip_sources(
            earned_tip,
            [f"{earned_source_col}: {_cell_for(earned_source_col, row_idx)}"] if earned_source_col else [],
        )
        gliss_tip = _append_tip_sources(
            gliss_tip,
            [f"{gliss_source_col}: {_cell_for(gliss_source_col, row_idx)}"] if gliss_source_col else [],
        )

        ecart_tip = _append_tip_sources(
            ecart_tip,
            [
                f"Earned: {_cell_for(earned_source_col, row_idx)}" if earned_source_col else "",
                f"Schedule: {schedule_week_cell}" if schedule_week_cell else "",
                f"Budgeted Units: {schedule_budget_cell}" if schedule_budget_cell else "",
            ],
            prefix="Sources",
        )
        impact_tip = _append_tip_sources(
            impact_tip,
            [
                f"Budgeted Units: {activity_budget_cell}" if activity_budget_cell else "",
                f"Root Budgeted Units: {root_budget_cell}" if root_budget_cell else "",
            ],
            prefix="Sources",
        )
        return {
            "planned_finish": planned_text,
            "planned_display": planned_display,
            "planned_tip": planned_tip,
            "forecast_finish": forecast_text,
            "forecast_display": forecast_display,
            "forecast_tip": forecast_tip,
            "schedule": schedule_val,
            "schedule_display": schedule_display,
            "schedule_tip": schedule_tip,
            "earned": earned_val,
            "earned_display": earned_display,
            "earned_tip": earned_tip,
            "ecart": ecart_val,
            "ecart_display": ecart_display,
            "ecart_tip": ecart_tip,
            "impact": impact_val,
            "impact_display": impact_display,
            "impact_tip": impact_tip,
            "glissement": gliss_val,
            "glissement_display": gliss_display,
            "glissement_tip": gliss_tip,
        }

    root: Dict | None = None
    stack: List[Dict] = []

    for _, r in df.iterrows():
        label = clean_label(r[label_col])
        if not label:
            continue
        lvl = space2lvl.get(r["_indent"], len(space2lvl))  # fallback = plus profond
        row_idx = int(r.get("_row_idx")) if "_row_idx" in r else None
        node = {"label": label, "level": int(lvl), "metrics": row_metrics(r, row_idx), "children": []}

        if not stack:
            root = node
            stack = [node]
            continue

        while stack and stack[-1]["level"] >= lvl:
            stack.pop()

        if not stack:
            # nouveau L1 → frère du root
            if root is None:
                root = node
                stack = [node]
            else:
                root.setdefault("children", []).append(node)
                stack = [root, node]
        else:
            stack[-1]["children"].append(node)
            stack.append(node)

    return root or {}

# ---------- Extraction (tous les tableaux) ----------
def extract_all_wbs(
    input_xlsx: str,
    schedule_lookup: Dict[str, Dict[str, Any]] | None = None,
    schedule_info: Dict[str, Any] | None = None,
    column_mapping: dict[str, dict[str, str]] | None = None,
) -> List[Dict]:
    wb = load_workbook(input_xlsx, data_only=True)
    results: List[Dict] = []

    if schedule_lookup is None or schedule_info is None:
        schedule_lookup, schedule_info = build_schedule_lookup(
            input_xlsx,
            column_mapping=column_mapping,
        )

    for ws in wb.worksheets:
        blocks = detect_all_blocks_with_left_extension(ws)
        for (r1, c1, r2, c2) in blocks:
            data = [[ws.cell(rr, cc).value for cc in range(c1, c2+1)] for rr in range(r1, r2+1)]
            if len(data) < 2:
                continue
            df = pd.DataFrame(data)
            df.columns = make_unique_columns([str(x).strip() if x is not None else "" for x in df.iloc[0].tolist()])
            df = df.iloc[1:, :].reset_index(drop=True)
            df, top_trim, left_trim, _, _ = trim_empty_border_with_offsets(df)
            source_meta = {
                "sheet": ws.title,
                "range": f"R{r1}C{c1}:R{r2}C{c2}",
                "data_row_start": r1 + 1 + top_trim,
                "data_col_start": c1 + left_trim,
            }

            # Filtre: doit contenir toutes les colonnes requises
            if not all(col in df.columns for col in REQUIRED_COLS):
                continue

            label_col = pick_label_col(df)
            tree = to_wbs_tree(
                df,
                label_col,
                schedule_lookup=schedule_lookup,
                schedule_info=schedule_info,
                source_meta=source_meta,
            )
            if tree:
                results.append({
                    "sheet": ws.title,
                    "range": f"R{r1}C{c1}:R{r2}C{c2}",
                    "wbs": tree
                })

    if results:
        return results

    # Fallback: use detected activity summary table
    summary = _load_detected_table(
        input_xlsx,
        "activity_summary",
        column_mapping=column_mapping,
    )
    if summary is None:
        return results
    df, meta, _ = summary
    label_col = "Activity ID" if "Activity ID" in df.columns else pick_label_col(df)
    tree = to_wbs_tree(
        df,
        label_col,
        schedule_lookup=schedule_lookup,
        schedule_info=schedule_info,
        source_meta=meta,
    )
    if tree:
        results.append({
            "sheet": meta["sheet"],
            "range": meta["range"],
            "wbs": tree,
        })
    return results

# ---------- CLI ----------
if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Build an array of WBS JSONs from all valid tables in Excel.")
    p.add_argument("input_xlsx", help="Path to Excel file, e.g., Book1.xlsx")
    p.add_argument("--out", default="wbs_all.json", help="Output JSON path")
    args = p.parse_args()

    all_wbs = extract_all_wbs(args.input_xlsx)
    Path(args.out).write_text(json.dumps(all_wbs, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"Saved: {args.out}  |  Tables matched: {len(all_wbs)}")
