from __future__ import annotations

import json
from datetime import date, datetime, timedelta
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook

DEMO_TEMPLATE_VERSION = 11
BASE_TEMPLATE_PATH = Path("artifacts") / "Chronoplan_Template.xlsx"
DEMO_TEMPLATE_DIR = Path(".streamlit") / "generated_demo"
GENERATED_TEMPLATE_PATH = DEMO_TEMPLATE_DIR / "Chronoplan_Template.xlsx"
META_PATH = DEMO_TEMPLATE_DIR / "Chronoplan_Template.meta.json"
PLANNED_WEEK_SHIFT_DAYS = 0
MIN_STEP_RATIO_BUDGETED = 0.05
MIN_STEP_RATIO_OTHER = 0.05


def _week_start(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _to_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _extract_week_header_cells(sheet) -> list[tuple[int, date]]:
    row = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if not row:
        return []
    headers: list[tuple[int, date]] = []
    for cell in row:
        header_date = _to_date(cell.value)
        if header_date is not None:
            headers.append((cell.col_idx, header_date))
    return headers


def _extract_week_headers(sheet) -> list[date]:
    row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not row:
        return []
    headers: list[date] = []
    for value in row:
        header_date = _to_date(value)
        if header_date is not None:
            headers.append(header_date)
    return headers


def _find_week_headers(workbook) -> list[date]:
    preferred = (
        "Ressource Assign. Budgeted",
        "Ressource Assign. Actual",
        "Ressource Assign. Remaining",
    )
    for name in preferred:
        if name in workbook.sheetnames:
            headers = _extract_week_headers(workbook[name])
            if headers:
                return headers

    for sheet in workbook.worksheets:
        headers = _extract_week_headers(sheet)
        if headers:
            return headers

    return []


def _calculate_delta(workbook, target_week: date) -> timedelta | None:
    # Anchor on the last weekly date of the Actual sheet (most intuitive for "current week" alignment).
    actual_name = "Ressource Assign. Actual"
    if actual_name in workbook.sheetnames:
        headers = _extract_week_headers(workbook[actual_name])
        if headers:
            last_actual = _week_start(headers[-1])
            return target_week - last_actual

    # Fallback: previous heuristic (pivot of any detected weekly headers)
    week_headers = _find_week_headers(workbook)
    if not week_headers:
        return None

    pivot_idx = len(week_headers) // 2
    pivot_header = week_headers[pivot_idx]
    pivot_target_week = _week_start(pivot_header - timedelta(days=PLANNED_WEEK_SHIFT_DAYS))
    return target_week - pivot_target_week


def _apply_delta(workbook, delta: timedelta) -> None:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, datetime):
                    cell.value = value + delta
                elif isinstance(value, date):
                    cell.value = value + delta


def _align_resource_headers(workbook, desired_start: date) -> None:
    """Align Budgeted/Actual/Remaining weekly headers.

    - Remaining starts at the current week (desired_start).
    - Actual ends one week before Remaining starts.
    - Budgeted follows Actual's cadence so weekly columns align.
    """

    def _align_sheet(sheet_name: str, start: date) -> tuple[date | None, timedelta, int]:
        if sheet_name not in workbook.sheetnames:
            return None, timedelta(days=7), 0
        sheet = workbook[sheet_name]
        header_cells = _extract_week_header_cells(sheet)
        if not header_cells:
            return None, timedelta(days=7), 0
        header_dates = [d for _, d in header_cells]
        if len(header_dates) > 1:
            step = header_dates[1] - header_dates[0]
            if step.days <= 0:
                step = timedelta(days=7)
        else:
            step = timedelta(days=7)

        for idx, (col_idx, _) in enumerate(header_cells):
            sheet.cell(row=1, column=col_idx).value = start + (step * idx)

        last_date = start + (step * (len(header_cells) - 1))
        return last_date, step, len(header_cells)

    # Align sheets:
    # - Actual and Budgeted share the same first week.
    # - Actual ends one week before Remaining starts.
    # - Remaining starts after Actual ends (by one step).
    # - Budgeted extends through the last Remaining week while keeping its start aligned with Actual.
    actual_sheet = "Ressource Assign. Actual"
    remaining_sheet = "Ressource Assign. Remaining"
    budget_sheet = "Ressource Assign. Budgeted"

    # Determine lengths and steps.
    remaining_cells = _extract_week_header_cells(workbook[remaining_sheet]) if remaining_sheet in workbook.sheetnames else []
    remaining_len = len(remaining_cells)
    remaining_step = timedelta(days=7)
    if remaining_len > 1:
        d = [d for _, d in remaining_cells]
        if len(d) > 1:
            s = d[1] - d[0]
            if s.days > 0:
                remaining_step = s

    budget_cells = _extract_week_header_cells(workbook[budget_sheet]) if budget_sheet in workbook.sheetnames else []
    budget_len = len(budget_cells)
    budget_step = timedelta(days=7)
    if budget_len > 1:
        d = [d for _, d in budget_cells]
        if len(d) > 1:
            s = d[1] - d[0]
            if s.days > 0:
                budget_step = s

    actual_cells = _extract_week_header_cells(workbook[actual_sheet]) if actual_sheet in workbook.sheetnames else []
    actual_len = len(actual_cells)
    actual_step = timedelta(days=7)
    if actual_len > 1:
        d = [d for _, d in actual_cells]
        if len(d) > 1:
            s = d[1] - d[0]
            if s.days > 0:
                actual_step = s

    # Remaining starts at the current week (desired_start).
    remaining_start = desired_start
    rem_step = timedelta(days=7)
    rem_len = remaining_len if remaining_len else actual_len or budget_len or 1
    if remaining_sheet in workbook.sheetnames:
        remaining_ws = workbook[remaining_sheet]
        start_col = remaining_cells[0][0] if remaining_cells else 6
        for idx in range(rem_len):
            remaining_ws.cell(row=1, column=start_col + idx).value = remaining_start + (rem_step * idx)
        remaining_len = rem_len

    # Actual ends at the current week; back-calculate its start.
    actual_step_used = actual_step if actual_step.days > 0 else timedelta(days=7)
    actual_len_eff = actual_len or rem_len or budget_len or 1
    actual_end = desired_start
    actual_start = actual_end - (actual_step_used * max(0, actual_len_eff - 1))
    if actual_sheet in workbook.sheetnames and actual_len_eff:
        start_col = actual_cells[0][0] if actual_cells else 6
        for idx in range(actual_len_eff):
            workbook[actual_sheet].cell(row=1, column=start_col + idx).value = actual_start + (actual_step_used * idx)

    # Budgeted start fixed to actual_start; extend to end at remaining_last.
    if budget_sheet in workbook.sheetnames:
        budget_ws = workbook[budget_sheet]
        # Preserve budget column count; only realign dates.
        target_len = budget_len
        start_col = budget_cells[0][0] if budget_cells else 6
        bud_step = budget_step if budget_step.days > 0 else timedelta(days=7)
        for idx in range(target_len):
            budget_ws.cell(row=1, column=start_col + idx).value = actual_start + (bud_step * idx)


def _normalize_resource_units(workbook, target_week: date) -> None:
    """Ensure Budgeted Units match across resource sheets and Remaining tops up to 100% at last week."""

    def _build_cumulative_values(week_cols: list[tuple[int, date]], target: float, min_step: float) -> list[float]:
        n = len(week_cols)
        if n == 0:
            return []
        values: list[float] = []
        prev = 0.0
        for idx in range(n):
            steps_left = n - idx - 1
            headroom = target - prev
            if steps_left == 0:
                next_val = max(prev, target)
            elif headroom <= 0:
                next_val = prev
            else:
                if headroom >= min_step * (steps_left + 1):
                    max_allowed = target - (min_step * steps_left)
                    next_val = min(max_allowed, prev + min_step)
                else:
                    next_val = prev + headroom / (steps_left + 1)
                next_val = min(next_val, target)
                if next_val < prev:
                    next_val = prev
            values.append(next_val)
            prev = next_val
        return values

    def _col_idx(sheet, header_name: str) -> int | None:
        row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not row:
            return None
        for idx, val in enumerate(row, start=1):
            if str(val).strip().lower() == header_name.lower():
                return idx
        return None

    def _week_cols(sheet):
        cols: list[tuple[int, date]] = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1), []):
            d = _to_date(cell.value)
            if d:
                cols.append((cell.col_idx, _week_start(d)))
        return cols

    budget_sheet_name = "Ressource Assign. Budgeted"
    actual_sheet_name = "Ressource Assign. Actual"
    remaining_sheet_name = "Ressource Assign. Remaining"

    if budget_sheet_name not in workbook.sheetnames:
        return
    budget_sheet = workbook[budget_sheet_name]
    budget_id_col = _col_idx(budget_sheet, "Activity ID")
    budget_units_col = _col_idx(budget_sheet, "Budgeted Units")
    budget_week_cols = _week_cols(budget_sheet)
    if budget_id_col is None or budget_units_col is None:
        return

    budget_map: dict[str, float] = {}
    for row in budget_sheet.iter_rows(min_row=2, values_only=False):
        activity = row[budget_id_col - 1].value
        if activity is None:
            continue
        try:
            units = float(row[budget_units_col - 1].value or 0)
        except Exception:
            units = 0.0
        budget_map[str(activity).strip()] = units

    def _apply_units(sheet_name: str) -> tuple[list[tuple[int, date]], dict[str, float]]:
        if sheet_name not in workbook.sheetnames:
            return [], {}
        sheet = workbook[sheet_name]
        id_col = _col_idx(sheet, "Activity ID")
        units_col = _col_idx(sheet, "Budgeted Units")
        weeks = _week_cols(sheet)
        values: dict[str, float] = {}
        if id_col is None or units_col is None:
            return weeks, values
        for row in sheet.iter_rows(min_row=2, values_only=False):
            activity = row[id_col - 1].value
            if activity is None:
                continue
            key = str(activity).strip()
            if key in budget_map:
                row[units_col - 1].value = budget_map[key]
                values[key] = budget_map[key]
        return weeks, values

    actual_weeks, actual_units_map = _apply_units(actual_sheet_name)
    remaining_weeks, remaining_units_map = _apply_units(remaining_sheet_name)

    def _build_cumulative_values(
        week_cols: list[tuple[int, date]], target: float, min_step: float
    ) -> list[float]:
        """Create a monotonic ramp that reaches target with minimum per-step increase when possible."""
        n = len(week_cols)
        if n == 0:
            return []
        values: list[float] = []
        prev = 0.0
        for idx in range(n):
            steps_left = n - idx - 1
            headroom = target - prev
            if steps_left == 0:
                next_val = max(prev, target)
            elif headroom <= 0:
                next_val = prev
            else:
                if headroom >= min_step * (steps_left + 1):
                    max_allowed = target - (min_step * steps_left)
                    next_val = min(max_allowed, prev + min_step)
                else:
                    next_val = prev + headroom / (steps_left + 1)
                next_val = min(next_val, target)
                if next_val < prev:
                    next_val = prev
            values.append(next_val)
            prev = next_val
        return values

    def _leaf_rows(sheet, id_col: int) -> list[int]:
        levels = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            raw_id = row[id_col - 1]
            if raw_id is None:
                levels.append(None)
                continue
            raw_id_str = str(raw_id)
            level = len(raw_id_str) - len(raw_id_str.lstrip(" "))
            levels.append(level)
        leaves: list[int] = []
        for i, lvl in enumerate(levels):
            if lvl is None:
                continue
            leaf = True
            for j in range(i + 1, len(levels)):
                nxt = levels[j]
                if nxt is None:
                    continue
                if nxt <= lvl:
                    break
                if nxt == lvl + 1:
                    leaf = False
                    break
            if leaf:
                leaves.append(i)
        return leaves

    def _apply_leaf_ramp(sheet_name: str, target_fn, ratio: float) -> None:
        if sheet_name not in workbook.sheetnames:
            return
        sheet = workbook[sheet_name]
        id_col = _col_idx(sheet, "Activity ID")
        units_col = _col_idx(sheet, "Budgeted Units")
        week_cols = sorted(_week_cols(sheet), key=lambda x: x[0])
        if id_col is None or units_col is None or not week_cols:
            return
        leaves = _leaf_rows(sheet, id_col)
        all_rows = list(sheet.iter_rows(min_row=2, values_only=False))
        for idx in leaves:
            if idx >= len(all_rows):
                continue
            row = all_rows[idx]
            activity = row[id_col - 1].value
            if activity is None:
                continue
            key = str(activity).strip()
            budget_units = budget_map.get(key, 0.0)
            target = target_fn(budget_units)
            min_step = (ratio * budget_units) if budget_units > 0 else 0.0
            ramp = _build_cumulative_values(week_cols, target, min_step)
            for (col_idx, _), val in zip(week_cols, ramp):
                row[col_idx - 1].value = val
            # ensure units column matches budget
            row[units_col - 1].value = budget_units

    def _enforce_cumulative(sheet_name: str, ratio: float) -> None:
        if sheet_name not in workbook.sheetnames:
            return
        sheet = workbook[sheet_name]
        id_col = _col_idx(sheet, "Activity ID")
        units_col = _col_idx(sheet, "Budgeted Units")
        week_cols = sorted(_week_cols(sheet), key=lambda x: x[0])
        if id_col is None or units_col is None or not week_cols:
            return
        for row in sheet.iter_rows(min_row=2, values_only=False):
            activity = row[id_col - 1].value
            if activity is None:
                continue
            key = str(activity).strip()
            budget_units = budget_map.get(key, 0.0)
            prev_val = 0.0
            for col_idx, _ in week_cols:
                raw = row[col_idx - 1].value
                try:
                    val = float(raw or 0.0)
                except Exception:
                    val = 0.0
                val = max(prev_val, val)  # cumulative non-decreasing
                if budget_units > 0 and val > prev_val:
                    min_step = (ratio * budget_units) if budget_units > 0 else 0.0
                    if val - prev_val < min_step:
                        val = min(prev_val + min_step, budget_units)
                val = min(val, budget_units)         # cap at budget
                row[col_idx - 1].value = val
                prev_val = val

    # Enforce cumulative monotonicity and cap for all resource sheets.
    _apply_leaf_ramp(budget_sheet_name, lambda b: b, MIN_STEP_RATIO_BUDGETED)
    _apply_leaf_ramp(actual_sheet_name, lambda b: max(0.0, b - max(400.0, 0.4 * b)), MIN_STEP_RATIO_OTHER)
    _apply_leaf_ramp(remaining_sheet_name, lambda b: max(0.0, min(b, max(400.0, 0.4 * b))), MIN_STEP_RATIO_OTHER)
    _enforce_cumulative(budget_sheet_name, MIN_STEP_RATIO_BUDGETED)
    _enforce_cumulative(actual_sheet_name, MIN_STEP_RATIO_OTHER)
    _enforce_cumulative(remaining_sheet_name, MIN_STEP_RATIO_OTHER)

    # Top up remaining cumulative values so Actual + Remaining ends at 100% of budget.
    if remaining_sheet_name in workbook.sheetnames:
        sheet = workbook[remaining_sheet_name]
        id_col = _col_idx(sheet, "Activity ID")
        units_col = _col_idx(sheet, "Budgeted Units")
        week_cols = sorted(_week_cols(sheet), key=lambda x: x[0])
        if id_col and units_col and week_cols:
            last_week_col_idx, _ = week_cols[-1]
            for row in sheet.iter_rows(min_row=2, values_only=False):
                activity = row[id_col - 1].value
                if activity is None:
                    continue
                key = str(activity).strip()
                budget_units = budget_map.get(key, 0.0)
                # Find actual cumulative at the last week date.
                actual_cum = 0.0
                actual_row_ref = None
                actual_last_col_idx = None
                if actual_sheet_name in workbook.sheetnames and actual_weeks:
                    actual_sheet = workbook[actual_sheet_name]
                    a_id_col = _col_idx(actual_sheet, "Activity ID")
                    a_units_col = _col_idx(actual_sheet, "Budgeted Units")
                    a_week_cols = sorted(_week_cols(actual_sheet), key=lambda x: x[0])
                    if a_id_col and a_units_col and a_week_cols:
                        try:
                            # locate matching row
                            for a_row in actual_sheet.iter_rows(min_row=2, values_only=False):
                                a_key = str(a_row[a_id_col - 1].value).strip() if a_row[a_id_col - 1].value is not None else ""
                                if a_key == key:
                                    actual_row_ref = a_row
                                    last_a_col_idx = a_week_cols[-1][0]
                                    cell_val = a_row[last_a_col_idx - 1].value
                                    actual_cum = float(cell_val or 0)
                                    break
                        except Exception:
                            actual_cum = 0.0

                # Enforce floor for Remaining (min 400 or 40% of budget, capped to budget).
                remaining_floor = min(budget_units, max(400.0, 0.4 * budget_units))
                max_actual_allowed = max(0.0, budget_units - remaining_floor)
                if actual_cum > max_actual_allowed:
                    actual_cum = max_actual_allowed
                    # Cap actual row values to new allowed max (keeps monotonic; re-enforced later).
                    if actual_row_ref is not None and last_a_col_idx is not None:
                        for col_idx, _ in a_week_cols:
                            try:
                                val = float(actual_row_ref[col_idx - 1].value or 0.0)
                            except Exception:
                                val = 0.0
                            actual_row_ref[col_idx - 1].value = min(val, max_actual_allowed)

                target_remaining = max(0.0, min(budget_units, budget_units - actual_cum))

                min_step = (MIN_STEP_RATIO_OTHER * budget_units) if budget_units > 0 else 0.0
                ramp = _build_cumulative_values(week_cols, target_remaining, min_step)
                for (col_idx, _), val in zip(week_cols, ramp):
                    row[col_idx - 1].value = val

                # Ensure units column matches budget
                row[units_col - 1].value = budget_units

    # Re-apply cumulative/min-step/cap after top-up to ensure consistency.
    _enforce_cumulative(actual_sheet_name, MIN_STEP_RATIO_OTHER)
    _enforce_cumulative(remaining_sheet_name, MIN_STEP_RATIO_OTHER)

    def _rollup_sheet(sheet_name: str) -> None:
        if sheet_name not in workbook.sheetnames:
            return
        sheet = workbook[sheet_name]
        id_col = _col_idx(sheet, "Activity ID")
        units_col = _col_idx(sheet, "Budgeted Units")
        week_cols = sorted(_week_cols(sheet), key=lambda x: x[0])
        if id_col is None or units_col is None or not week_cols:
            return

        rows_data = []
        for row in sheet.iter_rows(min_row=2, values_only=False):
            raw_id = row[id_col - 1].value
            if raw_id is None:
                continue
            raw_id_str = str(raw_id)
            level = len(raw_id_str) - len(raw_id_str.lstrip(" "))
            try:
                budget_val = float(row[units_col - 1].value or 0.0)
            except Exception:
                budget_val = 0.0
            week_vals = []
            for col_idx, _ in week_cols:
                try:
                    week_vals.append(float(row[col_idx - 1].value or 0.0))
                except Exception:
                    week_vals.append(0.0)
            rows_data.append((level, row, budget_val, week_vals))

        aggregates: list[tuple[float, list[float]]] = [None] * len(rows_data)  # type: ignore
        stack: list[tuple[int, float, list[float]]] = []

        for idx in range(len(rows_data) - 1, -1, -1):
            level, row_cells, budget_val, week_vals = rows_data[idx]
            agg_budget = 0.0
            agg_weeks = [0.0 for _ in week_vals]
            while stack and stack[-1][0] > level:
                _, child_budget, child_weeks = stack.pop()
                agg_budget += child_budget
                agg_weeks = [a + b for a, b in zip(agg_weeks, child_weeks)]
            if agg_budget == 0 and all(v == 0 for v in agg_weeks):
                agg_budget = budget_val
                agg_weeks = list(week_vals)
            aggregates[idx] = (agg_budget, agg_weeks)
            stack.append((level, agg_budget, agg_weeks))

        for idx, (_, row_cells, _, _) in enumerate(rows_data):
            agg_budget, agg_weeks = aggregates[idx]
            row_cells[units_col - 1].value = agg_budget
            for (col_idx, _), val in zip(week_cols, agg_weeks):
                row_cells[col_idx - 1].value = val

    # Roll up sums so parents equal the sum of direct children across all sheets.
    _rollup_sheet(budget_sheet_name)
    _rollup_sheet(actual_sheet_name)
    _rollup_sheet(remaining_sheet_name)

    # Keep Remaining headers aligned to the current week start.
    if remaining_weeks and remaining_sheet_name in workbook.sheetnames:
        sheet = workbook[remaining_sheet_name]
        sorted_weeks = sorted(remaining_weeks, key=lambda x: x[0])
        step = timedelta(days=7)
        if len(sorted_weeks) > 1:
            computed_step = sorted_weeks[1][1] - sorted_weeks[0][1]
            if computed_step.days > 0:
                step = computed_step
        start_date = target_week
        for offset, (col_idx, _) in enumerate(sorted_weeks):
            sheet.cell(row=1, column=col_idx).value = start_date + (step * offset)


def _ensure_readme_note(workbook) -> None:
    if "README" not in workbook.sheetnames:
        return
    note = "Remaining weekly columns should start at the current week."
    old_note = (
        "Remaining weekly columns should start at the last or penultimate "
        "planned week."
    )
    sheet = workbook["README"]
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(cell == note for cell in row):
            return
        if any(cell == old_note for cell in row):
            for col_idx, cell in enumerate(row, start=1):
                if cell == old_note:
                    sheet.cell(row=row_idx, column=col_idx).value = note
                    return
    last_row = 0
    for row_idx in range(sheet.max_row, 0, -1):
        row_values = [cell.value for cell in sheet[row_idx]]
        if any(value not in (None, "") for value in row_values):
            last_row = row_idx
            break
    target_row = last_row + 1 if last_row else 1
    if last_row:
        sheet.cell(row=target_row, column=1, value="")
        target_row += 1
    sheet.cell(row=target_row, column=1, value=note)


def _load_meta() -> dict | None:
    if not META_PATH.exists():
        return None
    try:
        return json.loads(META_PATH.read_text(encoding="utf-8"))
    except Exception:
        return None


def _save_meta(data: dict) -> None:
    DEMO_TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    META_PATH.write_text(json.dumps(data, indent=2), encoding="utf-8")


def _generate_live_template(target_week: date) -> Path:
    workbook = load_workbook(BASE_TEMPLATE_PATH)
    delta = _calculate_delta(workbook, target_week)
    print("[TEMPLATE] delta days:", None if delta is None else delta.days)
    if delta is None:
        raise ValueError("unable to locate weekly headers in the template")

    _apply_delta(workbook, delta)
    # Keep data intact: only shift dates; do not rewrite headers or values.
    DEMO_TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(GENERATED_TEMPLATE_PATH)
    return GENERATED_TEMPLATE_PATH


def ensure_demo_template() -> Path | None:
    if not BASE_TEMPLATE_PATH.exists():
        return None

    target_week = _week_start(date.today())
    base_mtime = BASE_TEMPLATE_PATH.stat().st_mtime
    meta = _load_meta()
    force = False
    try:
        force = str(st.query_params.get("force_template", "")) == "1"
    except Exception:
        force = False
    print("[TEMPLATE] base exists:", BASE_TEMPLATE_PATH.exists(), str(BASE_TEMPLATE_PATH))
    print("[TEMPLATE] target_week:", target_week.isoformat())
    print("[TEMPLATE] DEMO_TEMPLATE_VERSION:", DEMO_TEMPLATE_VERSION)
    print("[TEMPLATE] base_mtime:", base_mtime)
    print("[TEMPLATE] meta:", meta)
    print(
        "[TEMPLATE] generated exists:",
        GENERATED_TEMPLATE_PATH.exists(),
        str(GENERATED_TEMPLATE_PATH),
    )
    if GENERATED_TEMPLATE_PATH.exists():
        print("[TEMPLATE] generated mtime:", GENERATED_TEMPLATE_PATH.stat().st_mtime)
    print("[TEMPLATE] force refresh:", force)
    if (
        meta
        and meta.get("version") == DEMO_TEMPLATE_VERSION
        and meta.get("target_week") == target_week.isoformat()
        and isinstance(meta.get("base_mtime"), (int, float))
        and meta["base_mtime"] == base_mtime
        and GENERATED_TEMPLATE_PATH.exists()
        and not force
    ):
        return GENERATED_TEMPLATE_PATH

    try:
        path = _generate_live_template(target_week)
        _save_meta(
            {
                "version": DEMO_TEMPLATE_VERSION,
                "target_week": target_week.isoformat(),
                "base_mtime": base_mtime,
            }
        )
        return path
    except Exception as e:
        print("[TEMPLATE] generation failed:", repr(e))
        return None


def get_demo_template_path() -> Path | None:
    path = ensure_demo_template()
    if path and path.exists():
        return path
    if BASE_TEMPLATE_PATH.exists():
        return BASE_TEMPLATE_PATH
    return None


def demo_template_bytes() -> tuple[bytes, str] | tuple[None, None]:
    path = get_demo_template_path()
    if path and path.exists():
        return path.read_bytes(), path.name
    return None, None


def get_demo_template_debug() -> dict[str, object]:
    target_week = _week_start(date.today()).isoformat()
    meta = _load_meta()
    generated_exists = GENERATED_TEMPLATE_PATH.exists()
    base_exists = BASE_TEMPLATE_PATH.exists()
    generated_mtime = GENERATED_TEMPLATE_PATH.stat().st_mtime if generated_exists else None
    base_mtime = BASE_TEMPLATE_PATH.stat().st_mtime if base_exists else None
    served_path = (
        str(GENERATED_TEMPLATE_PATH)
        if generated_exists
        else (str(BASE_TEMPLATE_PATH) if base_exists else None)
    )
    return {
        "served_path": served_path,
        "generated_path": str(GENERATED_TEMPLATE_PATH),
        "generated_exists": generated_exists,
        "generated_mtime": generated_mtime,
        "base_path": str(BASE_TEMPLATE_PATH),
        "base_exists": base_exists,
        "base_mtime": base_mtime,
        "target_week": target_week,
        "meta": meta,
    }
